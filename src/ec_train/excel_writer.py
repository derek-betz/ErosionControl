"""Excel writer for EC Train learned features."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Iterable, Mapping

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

VERSION = "1.0"


@dataclass(slots=True)
class FeatureRow:
    """Row data for the EC Train workbook."""

    contract: str
    letting_date: str | None = None
    district: str | None = None
    route: str | None = None
    bidtabs_qty: float | None = None
    other_items: str | None = None
    spec_refs: str | None = None
    key_docs: Mapping[str, str] | None = None
    notes: str | None = None
    source_url: str | None = None


HEADERS = [
    "Contract",
    "LettingDate",
    "District",
    "Route",
    "BidTabsQty_205-12616",
    "Other_EC_Items",
    "Spec_Refs",
    "Key_Docs",
    "Notes/Findings",
    "Source_URL",
]


def autosize_columns(ws) -> None:
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(length + 2, 80)


def write_workbook(rows: Iterable[FeatureRow], output_dir: Path) -> Path:
    """Write the extracted features to an Excel workbook."""
    wb = Workbook()
    ws = wb.active
    ws.title = "EC Train"
    ws.append(HEADERS)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:J1"
    for row in rows:
        docs = "; ".join(f"{name}: {link}" for name, link in (row.key_docs or {}).items())
        ws.append(
            [
                row.contract,
                row.letting_date,
                row.district,
                row.route,
                row.bidtabs_qty,
                row.other_items,
                row.spec_refs,
                docs,
                row.notes,
                row.source_url,
            ]
        )
        if row.key_docs:
            cell = ws.cell(row=ws.max_row, column=8)
            first_link = next(iter(row.key_docs.values()))
            cell.value = docs or first_link
            cell.hyperlink = first_link
    autosize_columns(ws)
    wb.properties.title = "EC Train Learned Features"
    wb.properties.creator = "EC Train"
    wb.properties.version = VERSION
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"{date.today():%Y-%m-%d}_ec-train.xlsx"
    wb.save(output_path)
    return output_path


__all__ = ["FeatureRow", "HEADERS", "VERSION", "write_workbook"]
