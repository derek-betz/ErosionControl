"""Tests for Excel writer formatting."""

from pathlib import Path

from openpyxl import load_workbook

from ec_train.excel_writer import FeatureRow, HEADERS, write_workbook


def test_write_workbook_formats_sheet(tmp_path: Path):
    rows = [
        FeatureRow(
            contract="R-12345",
            letting_date="2024-02-01",
            district="Crawfordsville",
            route="I-65",
            bidtabs_qty=10.5,
            key_docs={"plan.pdf": "file:///tmp/plan.pdf"},
            notes="Found erosion details",
            source_url="https://erms.example/contract",
        )
    ]
    output = write_workbook(rows, tmp_path)
    assert output.exists()

    wb = load_workbook(output)
    ws = wb.active
    assert ws.title == "EC Train"
    assert ws.freeze_panes == "A2"
    assert ws.auto_filter.ref == "A1:J1"
    headers = [cell.value for cell in ws[1]]
    assert headers == HEADERS
    assert ws.cell(row=2, column=1).value == "R-12345"
